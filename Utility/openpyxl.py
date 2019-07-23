import openpyxl

class ExcelUtil:

    def setExcelFile(self, path):
        try:
            self.workbook = openpyxl.load_workbook(path)
        except:
            print('Failed to open the file')

    def getRowCount(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            return sheet.max_row
        except:
            print('Failed to get total row count')

    def getColCount(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            return sheet.max_column
        except:
            print('Failed to get total column count')

    def getCellData(self, sheetName, row_count, col_count):
        try:
            sheet = self.workbook[sheetName]
            return sheet.cell(row_count, col_count).value
        except:
            print("Failed to get the cell data")

